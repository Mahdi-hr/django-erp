"""
Export utilities for products and materials.
Generates Excel and PDF exports with professional designs.
"""
import io
import os
from decimal import Decimal
from django.http import HttpResponse
from django.utils import timezone


def _get_company_info():
    """Get company info from settings."""
    from apps.common.models import Setting
    logo_val = Setting.get_value('company_logo', '')
    logo_path = '' if logo_val.startswith('fa-') or not logo_val else logo_val
    return {
        'name': Setting.get_value('company_name', 'شرکت'),
        'address': Setting.get_value('company_address', ''),
        'phone': Setting.get_value('company_phone', ''),
        'logo': logo_path,
    }


def _jalali_now():
    """Get current Jalali datetime string."""
    try:
        import jdatetime
        return jdatetime.datetime.now().strftime('%Y/%m/%d %H:%M')
    except Exception:
        return timezone.now().strftime('%Y-%m-%d %H:%M')


def _format_currency(val):
    """Format number as Iranian currency."""
    try:
        return f'{int(val):,}'
    except (ValueError, TypeError):
        return '0'


def _get_logo_base64():
    """Get company logo as base64 data URI for PDF embedding."""
    from django.conf import settings
    company = _get_company_info()
    if not company['logo']:
        return ''
    logo_path = os.path.join(settings.MEDIA_ROOT, 'logos', company['logo'])
    if not os.path.exists(logo_path):
        return ''
    import base64
    ext = os.path.splitext(logo_path)[1].lower()
    mime = {'png': 'image/png', 'jpg': 'image/jpeg', 'jpeg': 'image/jpeg', 'svg': 'image/svg+xml'}.get(ext.lstrip('.'), 'image/png')
    with open(logo_path, 'rb') as f:
        data = base64.b64encode(f.read()).decode()
    return f'data:{mime};base64,{data}'


# ============================================================
# PRODUCTS EXPORTS
# ============================================================

def generate_products_excel(mode='employer'):
    """
    Generate Excel for products.
    mode='employer': all prices (cost, sale, wholesale, retail, dealer, export, special)
    mode='customer': only sale prices
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers

    company = _get_company_info()
    products = list(__import__('apps.products.models', fromlist=['Product']).Product.objects.filter(is_active=True))

    wb = Workbook()
    ws = wb.active
    ws.title = 'محصولات'

    if mode == 'employer':
        # Employer Excel - all prices
        cols = ['ردیف', 'کد', 'نام محصول', 'قیمت تمام‌شده', 'قیمت فروش', 'سود (%)',
                'قیمت عمده', 'قیمت جزئی', 'قیمت نمایندگی', 'قیمت صادراتی', 'قیمت ویژه']
        ws.merge_cells(f'A1:{chr(64+len(cols))}1')
        ws['A1'] = company['name']
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells(f'A2:{chr(64+len(cols))}2')
        ws['A2'] = 'گزارش جامع محصولات (برای کارفرما)'
        ws['A2'].font = Font(size=13, color='475569')
        ws['A2'].alignment = Alignment(horizontal='center')

        header_fill = PatternFill(start_color='1e293b', end_color='1e293b', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, h in enumerate(cols, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for i, p in enumerate(products, 1):
            row = i + 4
            data = [i, p.code, p.name, int(p.cost_price), int(p.sale_price),
                    f'{p.profit_percent}%',
                    int(p.wholesale_price), int(p.retail_price), int(p.dealer_price),
                    int(p.export_price), int(p.special_price)]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                if col in (4, 5, 7, 8, 9, 10, 11):
                    cell.number_format = '#,##0'

        col_widths = [8, 12, 25, 15, 15, 10, 15, 15, 15, 15, 15]
    else:
        # Customer Excel - only sale prices
        cols = ['ردیف', 'کد', 'نام محصول', 'قیمت فروش (ریال)']
        ws.merge_cells('A1:D1')
        ws['A1'] = company['name']
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:D2')
        ws['A2'] = 'لیست قیمت محصولات'
        ws['A2'].font = Font(size=13, color='475569')
        ws['A2'].alignment = Alignment(horizontal='center')

        header_fill = PatternFill(start_color='1e293b', end_color='1e293b', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, h in enumerate(cols, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for i, p in enumerate(products, 1):
            row = i + 4
            data = [i, p.code, p.name, int(p.sale_price)]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                if col == 4:
                    cell.number_format = '#,##0'

        col_widths = [8, 12, 30, 20]

    for i, w in enumerate(col_widths, 1):
        if i <= 26:
            ws.column_dimensions[chr(64 + i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    suffix = 'employer' if mode == 'employer' else 'customer'
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="products_{suffix}_{_jalali_now().replace("/", "-")}.xlsx"'
    return response


def generate_products_pdf(mode='employer'):
    """
    Generate PDF for products.
    mode='employer': invoice-like with all prices, logo, company info
    mode='customer': clean price list (only sale prices, no cost info)
    """
    from xhtml2pdf import pisa
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    company = _get_company_info()
    products = __import__('apps.products.models', fromlist=['Product']).Product.objects.filter(is_active=True)
    now = _jalali_now()
    logo_b64 = _get_logo_base64()

    # Register Tahoma font
    try:
        pdfmetrics.registerFont(TTFont('Tahoma', os.path.join('C:', os.sep, 'Windows', 'Fonts', 'tahoma.ttf')))
        pdfmetrics.registerFont(TTFont('Tahoma-Bold', os.path.join('C:', os.sep, 'Windows', 'Fonts', 'tahomabd.ttf')))
    except Exception:
        pass

    logo_html = f'<img src="{logo_b64}" style="height:60px;margin-bottom:8px;">' if logo_b64 else ''

    if mode == 'employer':
        title = 'گزارش جامع محصولات'
        subtitle = 'لیست تمام قیمت‌ها و حاشیه سود'
        headers = ['ردیف', 'کد', 'نام محصول', 'قیمت تمام‌شده', 'قیمت فروش', 'سود',
                   'عمده', 'جزئی', 'نمایندگی', 'صادراتی', 'ویژه']

        rows_html = ''
        for i, p in enumerate(products, 1):
            bg = '#f8fafc' if i % 2 == 0 else '#e2e8f0'
            rows_html += f'''<tr>
                <td style="background:{bg};">{i}</td>
                <td style="background:{bg};font-weight:600;">{p.code}</td>
                <td style="background:{bg};text-align:right;font-weight:500;">{p.name}</td>
                <td style="background:{bg};direction:ltr;">{_format_currency(p.cost_price)}</td>
                <td style="background:{bg};direction:ltr;font-weight:700;color:#059669;">{_format_currency(p.sale_price)}</td>
                <td style="background:{bg};direction:ltr;">{p.profit_percent}%</td>
                <td style="background:{bg};direction:ltr;">{_format_currency(p.wholesale_price)}</td>
                <td style="background:{bg};direction:ltr;">{_format_currency(p.retail_price)}</td>
                <td style="background:{bg};direction:ltr;">{_format_currency(p.dealer_price)}</td>
                <td style="background:{bg};direction:ltr;">{_format_currency(p.export_price)}</td>
                <td style="background:{bg};direction:ltr;">{_format_currency(p.special_price)}</td>
            </tr>'''
    else:
        title = 'لیست قیمت محصولات'
        subtitle = 'فقط قیمت‌های فروش'
        headers = ['ردیف', 'کد', 'نام محصول', 'قیمت فروش (ریال)']

        rows_html = ''
        for i, p in enumerate(products, 1):
            bg = '#f8fafc' if i % 2 == 0 else '#e2e8f0'
            rows_html += f'''<tr>
                <td style="background:{bg};">{i}</td>
                <td style="background:{bg};font-weight:600;">{p.code}</td>
                <td style="background:{bg};text-align:right;font-weight:500;">{p.name}</td>
                <td style="background:{bg};direction:ltr;font-weight:700;font-size:13px;color:#059669;">{_format_currency(p.sale_price)}</td>
            </tr>'''

    header_html = ''.join(f'<th>{h}</th>' for h in headers)

    html = f'''<!DOCTYPE html>
    <html dir="rtl" lang="fa">
    <head><meta charset="UTF-8">
    <style>
        @page {{ size: A4 landscape; margin: 12mm; }}
        body {{ font-family: Tahoma, sans-serif; direction: rtl; margin: 0; padding: 0; font-size: 11px; color: #1e293b; }}
        .header {{ text-align: center; border-bottom: 3px solid #1e293b; padding-bottom: 10px; margin-bottom: 12px; }}
        .company-name {{ font-size: 24px; font-weight: bold; color: #1e293b; margin: 0; }}
        .report-title {{ font-size: 15px; color: #475569; margin: 3px 0; font-weight: 600; }}
        .report-subtitle {{ font-size: 11px; color: #94a3b8; }}
        .meta {{ font-size: 9px; color: #94a3b8; margin-top: 4px; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 8px; }}
        th {{ background: #1e293b; color: #fff; padding: 7px 4px; font-size: 10px; text-align: center; border: 1px solid #334155; }}
        td {{ padding: 5px 4px; font-size: 9.5px; text-align: center; border: 1px solid #cbd5e1; }}
        .footer {{ text-align: center; font-size: 8px; color: #94a3b8; margin-top: 15px; border-top: 2px solid #e2e8f0; padding-top: 6px; }}
        .stamp {{ display: inline-block; border: 2px solid #cbd5e1; border-radius: 50%; width: 80px; height: 80px; text-align: center; line-height: 80px; color: #94a3b8; font-size: 9px; margin-top: 10px; }}
    </style></head>
    <body>
    <div class="header">
        {logo_html}
        <div class="company-name">{company['name']}</div>
        <div class="report-title">{title}</div>
        <div class="report-subtitle">{subtitle}</div>
        <div class="meta">تاریخ چاپ: {now} | تلفن: {company['phone']}</div>
    </div>
    <table>
        <thead><tr>{header_html}</tr></thead>
        <tbody>{rows_html}</tbody>
    </table>
    <div class="footer">
        {company['name']} — {company['address']} — {company['phone']}
        <br><div class="stamp">مهر و امضا</div>
    </div>
    </body></html>'''

    result = io.BytesIO()
    pisa_status = pisa.CreatePDF(io.BytesIO(html.encode('utf-8')), dest=result, encoding='utf-8')

    if pisa_status.err:
        return None

    suffix = 'employer' if mode == 'employer' else 'customer'
    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="products_{suffix}_{_jalali_now().replace("/", "-")}.pdf"'
    return response


# ============================================================
# MATERIALS EXPORTS
# ============================================================

def generate_materials_excel():
    """Generate Excel for materials with all details."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    company = _get_company_info()
    materials = list(__import__('apps.materials.models', fromlist=['Material']).Material.objects.filter(is_active=True).select_related('category'))

    wb = Workbook()
    ws = wb.active
    ws.title = 'مواد اولیه'

    cols = ['ردیف', 'کد', 'نام ماده', 'دسته‌بندی', 'واحد', 'قیمت خرید', 'موجودی', 'حداقل موجودی', 'وضعیت']
    ws.merge_cells(f'A1:{chr(64+len(cols))}1')
    ws['A1'] = company['name']
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells(f'A2:{chr(64+len(cols))}2')
    ws['A2'] = 'گزارش جامع مواد اولیه'
    ws['A2'].font = Font(size=13, color='475569')
    ws['A2'].alignment = Alignment(horizontal='center')

    header_fill = PatternFill(start_color='1e293b', end_color='1e293b', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(cols, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    status_map = {'sufficient': 'کافی', 'low': 'کمبود', 'out_of_stock': 'ناموجود'}
    for i, m in enumerate(materials, 1):
        row = i + 4
        status_fa = status_map.get(m.stock_status, 'کافی')
        data = [i, m.code, m.name, m.category.name, m.get_unit_display_fa(),
                int(m.purchase_price), float(m.current_stock), float(m.min_stock), status_fa]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            if col == 6:
                cell.number_format = '#,##0'
            if col == 9:
                if val == 'ناموجود':
                    cell.font = Font(color='FF0000', bold=True)
                elif val == 'کمبود':
                    cell.font = Font(color='FF8C00', bold=True)
                else:
                    cell.font = Font(color='008000', bold=True)

    col_widths = [8, 12, 25, 18, 10, 15, 12, 14, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="materials_{_jalali_now().replace("/", "-")}.xlsx"'
    return response


def generate_materials_pdf(mode='employer'):
    """
    Generate PDF for materials.
    mode='employer': invoice-like with all details, logo, company info
    mode='customer': price list only
    """
    from xhtml2pdf import pisa
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    company = _get_company_info()
    materials = __import__('apps.materials.models', fromlist=['Material']).Material.objects.filter(is_active=True).select_related('category')
    now = _jalali_now()
    logo_b64 = _get_logo_base64()

    try:
        pdfmetrics.registerFont(TTFont('Tahoma', os.path.join('C:', os.sep, 'Windows', 'Fonts', 'tahoma.ttf')))
        pdfmetrics.registerFont(TTFont('Tahoma-Bold', os.path.join('C:', os.sep, 'Windows', 'Fonts', 'tahomabd.ttf')))
    except Exception:
        pass

    logo_html = f'<img src="{logo_b64}" style="height:60px;margin-bottom:8px;">' if logo_b64 else ''

    if mode == 'employer':
        title = 'گزارش جامع مواد اولیه'
        subtitle = 'لیست کامل مواد اولیه و موجودی انبار'
        headers = ['ردیف', 'کد', 'نام ماده', 'دسته‌بندی', 'واحد', 'قیمت خرید', 'موجودی', 'حداقل', 'وضعیت']

        rows_html = ''
        for i, m in enumerate(materials, 1):
            bg = '#f8fafc' if i % 2 == 0 else '#e2e8f0'
            status_fa = {'sufficient': 'کافی', 'low': 'کمبود', 'out_of_stock': 'ناموجود'}.get(m.stock_status, 'کافی')
            status_color = {'sufficient': '#059669', 'low': '#d97706', 'out_of_stock': '#dc2626'}.get(m.stock_status, '#059669')
            rows_html += f'''<tr>
                <td style="background:{bg};">{i}</td>
                <td style="background:{bg};font-weight:600;">{m.code}</td>
                <td style="background:{bg};text-align:right;font-weight:500;">{m.name}</td>
                <td style="background:{bg};">{m.category.name}</td>
                <td style="background:{bg};">{m.get_unit_display_fa()}</td>
                <td style="background:{bg};direction:ltr;">{_format_currency(m.purchase_price)}</td>
                <td style="background:{bg};direction:ltr;font-weight:700;">{float(m.current_stock)}</td>
                <td style="background:{bg};direction:ltr;">{float(m.min_stock)}</td>
                <td style="background:{bg};color:{status_color};font-weight:700;">{status_fa}</td>
            </tr>'''
    else:
        title = 'لیست قیمت مواد اولیه'
        subtitle = 'فقط قیمت‌ها'
        headers = ['ردیف', 'کد', 'نام ماده', 'واحد', 'قیمت خرید (ریال)']

        rows_html = ''
        for i, m in enumerate(materials, 1):
            bg = '#f8fafc' if i % 2 == 0 else '#e2e8f0'
            rows_html += f'''<tr>
                <td style="background:{bg};">{i}</td>
                <td style="background:{bg};font-weight:600;">{m.code}</td>
                <td style="background:{bg};text-align:right;font-weight:500;">{m.name}</td>
                <td style="background:{bg};">{m.get_unit_display_fa()}</td>
                <td style="background:{bg};direction:ltr;font-weight:700;font-size:13px;color:#059669;">{_format_currency(m.purchase_price)}</td>
            </tr>'''

    header_html = ''.join(f'<th>{h}</th>' for h in headers)

    html = f'''<!DOCTYPE html>
    <html dir="rtl" lang="fa">
    <head><meta charset="UTF-8">
    <style>
        @page {{ size: A4 landscape; margin: 12mm; }}
        body {{ font-family: Tahoma, sans-serif; direction: rtl; margin: 0; padding: 0; font-size: 11px; color: #1e293b; }}
        .header {{ text-align: center; border-bottom: 3px solid #1e293b; padding-bottom: 10px; margin-bottom: 12px; }}
        .company-name {{ font-size: 24px; font-weight: bold; color: #1e293b; margin: 0; }}
        .report-title {{ font-size: 15px; color: #475569; margin: 3px 0; font-weight: 600; }}
        .report-subtitle {{ font-size: 11px; color: #94a3b8; }}
        .meta {{ font-size: 9px; color: #94a3b8; margin-top: 4px; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 8px; }}
        th {{ background: #1e293b; color: #fff; padding: 7px 4px; font-size: 10px; text-align: center; border: 1px solid #334155; }}
        td {{ padding: 5px 4px; font-size: 9.5px; text-align: center; border: 1px solid #cbd5e1; }}
        .footer {{ text-align: center; font-size: 8px; color: #94a3b8; margin-top: 15px; border-top: 2px solid #e2e8f0; padding-top: 6px; }}
        .stamp {{ display: inline-block; border: 2px solid #cbd5e1; border-radius: 50%; width: 80px; height: 80px; text-align: center; line-height: 80px; color: #94a3b8; font-size: 9px; margin-top: 10px; }}
    </style></head>
    <body>
    <div class="header">
        {logo_html}
        <div class="company-name">{company['name']}</div>
        <div class="report-title">{title}</div>
        <div class="report-subtitle">{subtitle}</div>
        <div class="meta">تاریخ چاپ: {now} | تلفن: {company['phone']}</div>
    </div>
    <table>
        <thead><tr>{header_html}</tr></thead>
        <tbody>{rows_html}</tbody>
    </table>
    <div class="footer">
        {company['name']} — {company['address']} — {company['phone']}
        <br><div class="stamp">مهر و امضا</div>
    </div>
    </body></html>'''

    result = io.BytesIO()
    pisa_status = pisa.CreatePDF(io.BytesIO(html.encode('utf-8')), dest=result, encoding='utf-8')

    if pisa_status.err:
        return None

    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="materials_{_jalali_now().replace("/", "-")}.pdf"'
    return response
